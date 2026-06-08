import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "reports" / "All.xlsx"
OUTPUT = ROOT / "public" / "sharq-candidates.json"

COMPLEXES = {
    "34A": {"size": "large", "buildings": 10},
    "45D": {"size": "large", "buildings": 10},
    "46A": {"size": "large", "buildings": 10},
    "43D": {"size": "small", "buildings": 8},
    "43F": {"size": "small", "buildings": 8},
    "44A": {"size": "small", "buildings": 8},
    "44B": {"size": "small", "buildings": 8},
}

LARGE_TYPES = {
    1: "tip1",
    2: "tip2",
    3: "tip2",
    4: "tip1",
    5: "tip3_large",
    6: "tip1",
    7: "tip2",
    8: "tip2",
    9: "tip1",
    10: "tip3_large",
}

SMALL_TYPES = {
    1: "tip1",
    2: "tip2",
    3: "tip1",
    4: "tip3_small",
    5: "tip1",
    6: "tip2",
    7: "tip1",
    8: "tip3_small",
}

TYPE_CONFIGS = {
    "tip1": {
        "label": "TIP I - 12-storey",
        "floors": range(3, 13),
        "per_floor": 5,
        "default_start": 11,
    },
    "tip1_4": {
        "label": "12-storey, 4 flats/storey",
        "floors": range(3, 13),
        "per_floor": 4,
        "default_start": 5,
    },
    "tip2": {
        "label": "TIP II - 9-storey",
        "floors": range(3, 10),
        "per_floor": 6,
        "default_start": 13,
    },
    "tip3_large": {
        "label": "TIP III - 9-storey",
        "floors": range(3, 10),
        "per_floor": 6,
        "default_start": 13,
    },
    "tip3_small": {
        "label": "TIP III - 7-storey",
        "floors": range(3, 8),
        "per_floor": 6,
        "default_start": 13,
    },
}

TYPE_OVERRIDES = {
    ("46A", 4): "tip1_4",
}


def ranges(numbers):
    values = sorted(numbers)
    if not values:
        return ""

    result = []
    start = previous = values[0]
    for number in values[1:]:
        if number == previous + 1:
            previous = number
            continue
        result.append(str(start) if start == previous else f"{start}-{previous}")
        start = previous = number

    result.append(str(start) if start == previous else f"{start}-{previous}")
    return ", ".join(result)


def parse_flat_name(value):
    match = re.match(r"^([^/]+)/([^/]+)/([^/]+)$", str(value or "").strip())
    if not match:
        return None
    block, building, flat = match.groups()
    try:
        return block, int(building), int(flat)
    except ValueError:
        return None


def infer_start(observed, config):
    if not observed:
        return config["default_start"]

    floors = list(config["floors"])
    min_floor = min(floors)
    per_floor = config["per_floor"]
    scores = []

    for start in range(1, 25):
        inventory = {
            start + (floor - min_floor) * per_floor + slot
            for floor in floors
            for slot in range(per_floor)
        }
        contained = sum(1 for _, flat in observed if flat in inventory)
        scores.append((contained, -abs(start - config["default_start"]), start))

    scores.sort(reverse=True)
    return scores[0][2]


def read_workbook():
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    observed = defaultdict(list)
    statuses = defaultdict(lambda: {"sold": set(), "onSale": set()})

    for sheet_name, status in [("All", "sold"), ("OnSale", "onSale")]:
        worksheet = workbook[sheet_name]
        headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
        indexes = {header: index for index, header in enumerate(headers)}

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            parsed = parse_flat_name(row[indexes["name"]])
            if not parsed:
                continue

            block, building, flat = parsed
            if block not in COMPLEXES:
                continue

            unit_floor = row[indexes["unit_floor"]]
            key = (block, building)
            if unit_floor is not None:
                observed[key].append((int(unit_floor), flat))
            statuses[key][status].add(flat)

    return observed, statuses


def build_candidates():
    observed, statuses = read_workbook()
    result = {}

    for block, metadata in COMPLEXES.items():
        type_map = LARGE_TYPES if metadata["size"] == "large" else SMALL_TYPES
        rows = []
        total = 0

        for building in range(1, metadata["buildings"] + 1):
            type_id = TYPE_OVERRIDES.get((block, building), type_map[building])
            config = TYPE_CONFIGS[type_id]
            floors = list(config["floors"])
            min_floor = min(floors)
            per_floor = config["per_floor"]
            start = infer_start(observed[(block, building)], config)
            inventory = {
                start + (floor - min_floor) * per_floor + slot
                for floor in floors
                for slot in range(per_floor)
            }
            sold = statuses[(block, building)]["sold"] & inventory
            on_sale = statuses[(block, building)]["onSale"] & inventory
            later = sorted(inventory - sold - on_sale)

            total += len(later)
            rows.append(
                {
                    "building": str(building),
                    "type": config["label"],
                    "expected": len(inventory),
                    "sold": len(sold),
                    "onSale": len(on_sale),
                    "later": len(later),
                    "flats": ranges(later),
                    "message": "All expected residential flats are sold or on sale."
                    if not later
                    else "",
                }
            )

        size_label = metadata["size"].capitalize()
        result[block] = {
            "complex": f"{size_label} complex - {metadata['buildings']} buildings",
            "total": total,
            "sections": rows,
        }

    return result


def main():
    if not WORKBOOK.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK}")

    data = build_candidates()
    OUTPUT.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    totals = Counter()
    for block, block_data in data.items():
        totals[block] = block_data["total"]
    print(f"Wrote {OUTPUT}")
    print("Later-auction candidate totals:")
    for block in sorted(totals):
        print(f"  {block}: {totals[block]}")


if __name__ == "__main__":
    main()
