import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "reports" / "All.xlsx"
OUTPUT = ROOT / "public" / "sharq-onsale-flats.json"
SHEET_NAME = "OnSale"

COMPLEXES = {"34A", "45D", "46A", "43D", "43F", "44A", "44B"}


def parse_flat_code(value):
    match = re.match(r"^([0-9]{2}[A-Z])/(\d{1,2})/(\d{1,3})$", str(value or "").strip().upper())
    if not match:
        return None
    lot_id, building, flat = match.groups()
    return lot_id, building, int(flat)


def read_onsale_flats():
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")

    worksheet = workbook[SHEET_NAME]
    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    indexes = {header: index for index, header in enumerate(headers) if header}
    required = ["name", "lot_number", "price", "unit_floor", "building_floors", "total_area", "rooms"]
    missing = [header for header in required if header not in indexes]
    if missing:
        raise SystemExit(f"Missing required OnSale columns: {', '.join(missing)}")

    flats = {}

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        parsed = parse_flat_code(row[indexes["name"]])
        if not parsed:
            continue

        lot_id, building, flat = parsed
        if lot_id not in COMPLEXES:
            continue

        code = f"{lot_id}/{building}/{flat}"
        flats[code] = {
            "code": code,
            "lotId": lot_id,
            "building": building,
            "flat": flat,
            "lotNumber": row[indexes["lot_number"]],
            "price": row[indexes.get("price")],
            "deposit": row[indexes.get("deposit")] if "deposit" in indexes else None,
            "applications": row[indexes.get("applications")] if "applications" in indexes else None,
            "status": row[indexes.get("status")] if "status" in indexes else "On sale",
            "auctionEnd": row[indexes.get("auction_end")] if "auction_end" in indexes else None,
            "unitFloor": row[indexes["unit_floor"]],
            "buildingFloors": row[indexes["building_floors"]],
            "totalArea": row[indexes["total_area"]],
            "rooms": row[indexes["rooms"]],
            "completionTerm": row[indexes.get("completion_term")] if "completion_term" in indexes else None,
            "pricePerSqm": row[indexes.get("price_per_sqm")] if "price_per_sqm" in indexes else None,
        }

    return flats


def main():
    if not WORKBOOK.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK}")

    flats = read_onsale_flats()
    payload = {
        "source": str(WORKBOOK.relative_to(ROOT)).replace("\\", "/"),
        "sheet": SHEET_NAME,
        "count": len(flats),
        "flats": dict(sorted(flats.items())),
    }
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    print(f"On-sale flats: {len(flats)}")


if __name__ == "__main__":
    main()
